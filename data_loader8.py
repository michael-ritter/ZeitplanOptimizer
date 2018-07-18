import re
import numpy as np
from openpyxl import load_workbook
import argparse
import os

# path parameter
folder=os.path.abspath(".")
currentpath=folder

if folder[-6:] == "python":
    folder = os.path.dirname(folder)
    currentpath = os.path.join(folder,"python")
    

#import sys

#sys.path.append(currentpath)

# In[157]:

def look_up(ws,value=""):

    """

    find the coordinate of the first cell 

    that contains the substring value

    

    Args:

        ws: the worksheet

        value: value to look for

    

    Return:

        cell: first encountered cell containing value

    """

    value = value.strip()

    for row in ws.iter_rows():

        for cell in row:

            if str(cell.value).strip() == value:

                return cell

def is_none(cell):

    """return True if a cell can be considered empty"""

    if type(cell) is str:

        text=cell

    else:

        text=cell.value

    return (text=="None" or text is None or

            text=="" or text == 0)

def to_float(cell,default=0):

    """Try to read a cell as a float number, 

    return 'default' otherwise"""

    try:

        return float(cell.value)

    except:

        return default

def to_int(cell,default=0):

    """Try to read a cell as an integer, 

    return 'default' otherwise"""

    try:

        return int(cell.value)

    except:

        return default

def to_str(cell,default=""):

    """Try to read a cell as a string, 

    return 'default' otherwise"""

    if is_none(cell):

        return default

    else:

        s=str(cell.value).strip()

        for i in [("ü","ue"),("ö","oe"),("ä","ae"),("Ü","UE"),("Ö","OE"),("Ä","AE"),("ß","ss")]:

            s = s.replace(i[0],i[1])

        return s

    

def to_list(cell,sep='[,;]'):

    """Try to read a cell as a list, 

    

    Args: 

        sep: a regex for the element separator in the cell"""

    if type(cell) is str:

        text=cell

    else:

        text=to_str(cell)

    res=re.split(sep, text)

    res=[i.strip() for i in res]

    res=[i for i in res if not is_none(i)]

    return res

def to_list_2(cell,sep=','):

    """reads a cell containing a string such as

    'a1,a2,a3(b31,b32),a4(b41,b42,b43),a5,a6' """

    if not type(cell)==str:

        cell=to_str(cell)

    temp=[[j.strip("[, ]") for j in i.split("(")]

          for i in cell.split(")")]

    res=[]

    for i in temp:

        names1=[j.strip() for j in i[0].split(sep)]

        if len(i)>1:

            names2=[j.strip() for j in i[1].split(sep)]

            res.extend([(j,None) for j in names1[:-1]])

            res.append((names1[-1],names2))

        else:

            res.extend([(j,None) for j in names1])

    res=[i for i in res if not is_none(i[0])]

    return res

def separate_faecher(hauptfaecher, fachlist):

    """return a list of courses for the optimization model

    

    Args:

        hauptfaecher: <list> courses considered as main course ("Hauptfach")

        fachlist: <list of set> (courses and their properties)

    

    Return:

        res: <list of set> [(course name, 

                            splitted in .. groups,

                            need .. tandem teachers,

                            dauer in stunden)]

    """

    res=[]

    for f in fachlist:

        name= "Hauptfach" if f[1] in hauptfaecher else f[1]

        tandem=[f[2]>f[3],f[3]>f[4] and f[3]>0,f[4]>0]

        tandem=[i[0] for i in enumerate(tandem) if i[1]]

        if f[7]=="kann":

            for i in np.arange(2,f[6]+1):

                res.append((name,i,0,f[5]))

            for t in tandem:

                res.append((name,1,t,f[5]))

        else:

            for t in tandem:

                res.append((name,f[6],t,f[5]))

        

    return res

    

def add_gleichgultig(gg_list,gg_set,stunden):

    index=[i[0] for i in enumerate(gg_list) if i[1][0]==gg_set]

    if len(index)>0:

        gg_list[index[0]][1]+=stunden

    else:

        gg_list.append([gg_set,stunden])

def write_list(l):

    if type(l) in [float,int,bool]:

        return "%s"%l

    if type(l)==str:

        return "'%s'"%l

    s="["

    as_number = sum([type(i) not in (int,float) for i in l])==0

    if as_number:

        for i in l:

            s+='%s '%i

    else:

        for i in l:

            s+='"%s" '%i

    s+="]"

    return s

      

def write_tuple(l):

    s=""

    for i in l:

        if type(i) in (int,float):

            s="%s %s"%(s,i)

        else:

            s="%s '%s'"%(s,i)

    s="(%s)"%s

    return s

    

def write_dict(dic, tuple_key=None):

    if tuple_key is not None:

        temp='\n'.join(["%s %s"%(write_tuple(i[0]),

                        write_list(i[1])) for i in dic.items()])

    else:

        temp='\n'.join(["('%s') %s"%(i[0],write_list(i[1])) 

                        for i in dic.items()])

    temp="[%s]"%temp

    return temp

def write_list_list(vals, sep="\n ", bracket=True):

    if bracket:

        res=sep.join([write_list(i) for i in vals])

    else:

        res=sep.join([write_list(i)[1:-1] for i in vals])

    res="[ %s ]"%res

    return res

def write_list3(vals, sep="\n "):

    res=sep.join(write_list_list(i, sep=" ") for i in vals)

    res="[ %s ]"%res

    return res

# In[165]:

class reader(object):

    """An object that reads an excel file containing school timetable 

    requirements and translate it in a mosel readable .dat file

    """

    

    def __init__(self,source="source.xlsx",target_file="target.dat"):

        """Initialize the reader object

        Args:
            source: <str> path of the excel data file
            target_file: <str> path of the new mosel .dat file
        """

        self.data = load_workbook(source)
        self.message = []
        self.target_file=target_file
        
        self.check_worksheet()
        self.import_faecher()
        self.import_faecherstunden()
        self.import_lehrer()
        self.import_klassen_lehrer()
        self.import_klassen_tage()
        self.import_raume()
        self.import_ubergreifend()
        self.import_vorgaben()
        self.import_lehrerverfugbarkeiten()
        self.import_gleichzeitig()
        
        self.check_fach_taught()
        self.check_vorgaben()
        self.check_vorgaben_ml()
        self.check_faecher_raw()
        self.check_raume()
        
        
    def iter_row_from(self,sheet="Lehrpläne",start="Klasse",cols=[],offset=(0,0)):

        ws=self.data[sheet]

        start=look_up(ws,start)

        res=list(ws.iter_rows(min_col=start.col_idx+offset[0],

                 min_row=start.row+1+offset[1]))

        if cols != []:

            index=[look_up(ws,i).col_idx-start.col_idx for i in cols]

            res=[[i[j] for j in index] for i in res]

        return res

    def add_fach(self,faecher,fach_list,check_geteilt=None):

        temp=[]

        if "Deutsch" in fach_list:

            fach="Hauptfach"

            temp.extend([i for i in self.fach_list if 

                         (fach+"__") in i or fach==i ])

        fach_list=[f for f in fach_list if 

                   (f not in self.hauptfaecher)]

        for f in fach_list:

            temp.extend([i for i in self.fach_list if 

                     (f+"__") in i or f==i ])

        if check_geteilt == ">1":

            temp=[f for f in temp if 

                  self.geteilt[self.fach_dict[f]] > 1]

        elif check_geteilt is not None:

            temp=[f for f in temp if 

                  self.geteilt[self.fach_dict[f]] == check_geteilt]

        faecher.extend(temp)    

    def check_in_lehrer(self,lehrer_list,sheet=""):

        """Check that lehrer_list does not contain any unknown teacher

        Return the set of unknown teacher and the list of known ones"""

        if not "lehrer_set" in self.__dict__.keys():

            self.lehrer_set=set(self.lehrer_list)

        lehrer_set=set(lehrer_list)

        error = lehrer_set.difference(self.lehrer_set)

        if len(error)>0:

            self.message.append("Warning: im '%s', "%sheet+

                "sind die Lehrer %r unbekannt"%error)

        return error,list(lehrer_set.difference(error))

    def check_in_klassen(self,klassen_list,sheet=""):

        """Check that klassen_list does not contain any unknown class

        Return the set of unknown classes and the list of known ones"""

        if not "klassen_set" in self.__dict__.keys():

            self.klassen_set=set(self.klassen_list)

        klassen_set=set(klassen_list)

        error = klassen_set.difference(self.klassen_set)

        if len(error)>0:

            self.message.append("Warning: im '%s', "%sheet+

                "sind die Klassen %r unbekannt"%error)

        return error,list(klassen_set.difference(error))

    def check_in_faecher(self,fach_list,sheet=""):

        """Check that faecher_list does not contain any unknown course

        Return the set of unknown courses and the list of known ones"""

        fach_set=set(fach_list)

        error = fach_set.difference(self.fach_set)

        if len(error)>0:

            self.message.append("Warning: im '%s', "%sheet+

                "sind die faecher %r unbekannt"%error)

        return error,list(fach_set.difference(error))

    def check_all_klassen(self,klassen_list,sheet=""):

        """Check that klassen_list contains all known classes

        Return a list of known classes that were not mentioned"""

        if not "klassen_set" in self.__dict__.keys():

            self.klassen_set=set(self.klassen_list)

        error = self.klassen_set.difference(set(klassen_list))

        if len(error)>0:

            self.message.append("Warning: im '%s', "%sheet+

                "fehlen die Klassen %r"%error)

        return error

    def check_zeiten(self,zeiten,sheet="",default=0):

        """Check that zeiten is in the correct format and returns it

        

        Args:

            zeiten: <list of list> [[Mon_1,Di_1,Mi_1,Do_1,Fr_1],[Mon_2, ...], ...]

        

        Return:

            res: <list> [Mon_1,Mon_2,..,Di_1,Di_2,..,..]

        """

        zeiten=list(zip(*zeiten))

        res=[]

        for i in enumerate(zeiten):

            diff= self.woche[i[0]] - len(i[1])

            temp=i[1]

            if diff>0:

                s="Warning: im %s, Tag %d, fehlt es Stunden"%(sheet,i[0])

                self.message.append(s)

                i[1].extend([default]*(self.woche[i[0]]-len(i[1])) )

                temp=i[1]

            if diff<0:

                temp=i[1][:diff]

            res.extend(temp)

        return res

    def check_worksheet(self):

        ws_list=["Lehrerübersicht","Übergreifende Fächer","Lehrpläne",

                "Stundenzahl pro Tag","Klassenlehrer","Räume",

                "Feste Vorgaben","Lehrerverfügbarkeiten","Gleichzeitige Fächer"]

        

        gegeben = [i.title for i in self.data.worksheets]

        for ws in ws_list:

            if not ws in gegeben:

                self.message.append("Missing worksheet: '%s'"%ws)
  
    def check_fach_taught(self):
        """Check if every course has some teacher to be taught"""
        hours_due = {f:h for f,h in zip(self.fach_list,np.sum(self.stunden,0))}
        hours_possible = {f:0 for f in self.fach_list}
        
        for hours,faecher in zip(self.arbeitzeit,self.lehrer_faecher):
            for f in faecher:
                hours_possible[f]+=hours
        for f in self.fach_list:
            p,d=hours_possible[f],hours_due[f]
            if p<d:
                self.message.append("Warning: Not enough teacher available to teach subject '%s'"%f+
                                    "(%.f hours due, %.f hours possible)"%(d,p))
        return None                                                                               

    def check_vorgaben(self):
        """Pruft ,dass die Vorgaben andere Eingabe nicht widersprechen"""
        for k in self.vorgaben.keys():
            fach = self.vorgaben[k][0]
            lehr=[l for l in self.lehrer_list if fach in self.lehrer_faecher[self.lehrer_dict[l]]]
            for z in self.vorgaben_zeiten[k]:
                lehr2=[]
                
                #Prüft, dass ein Lehrer an diesen Zeitpunkt frei ist.
                for l in lehr:
                    if self.arbeitzeit[self.lehrer_dict[l]]==0:
                        continue
                    if l in self.lehrer_verfugbar:
                        if z in self.lehrer_verfugbar[l]:
                            lehr2.append(l)
                    else:
                        lehr2.append(l)
                if len(lehr2)==0:
                    s="Kein Lehrer ist für die Vorgabe '%s'"%fach
                    s+=" in Klasse %s zum zeitpunkt %s verfugbar"%(k[0],z)
                    self.message.append(s)
                    
                # Prüft, dass ein Raum dafür frei ist
                for rf,rv in zip(self.raum_faecher,self.raum_verfugbar):
                    if fach in rf and rv[z-1]<1:
                        s="Kein freier Raum für die Vorgabe"
                        s+=" '%s' in Klasse %s zum Zeitpunkt %s"%(fach,k[0],z)
                        self.message.append(s)

    def check_vorgaben_ml(self):
        """Pruft ,dass die Vorgaben andere Eingabe nicht widersprechen"""
        for k in self.vorgaben_mit_lehrer.keys():
            fach = self.vorgaben_mit_lehrer[k][0]
            for z in self.vorgaben_ml_zeiten[k]:
            
                #Prüft, dass ein Lehrer an diesen Zeitpunkt frei ist.
                verfug=True
                if self.arbeitzeit[self.lehrer_dict[k[1]]]==0:
                    verfug=False
                if k[1] in self.lehrer_verfugbar:
                    if z not in self.lehrer_verfugbar[k[1]]:
                        verfug=False
                if not verfug:
                    s="Lehrer '%s' ist für die Vorgabe '%s'"%(k[1],fach)
                    s+=" in Klasse %s zum zeitpunkt %s verfugbar "%(k[0],z)
                    self.message.append(s)
                    
                # Prüft, dass ein Raum dafür frei ist
                for rf,rv in zip(self.raum_faecher,self.raum_verfugbar):
                    if fach in rf and rv[z]!=1:
                        s="Kein freier Raum für die Vorgabe"
                        s+=" '%s' in Klasse %s zum Zeitpunkt %s"%(fach,k[0],z)
                        self.message.append(s)
            
    def check_faecher_raw(self):
        """The number of required hours for a subject should be a multiple of its duration"""
        for KlassenFaecher in self.faecher_raw:
            for f in KlassenFaecher:
                columns=[0,0,"Unterrichtstunden","mit Tandem Lehrer", "mit 2 Tandem Lehrer"]
                for i in [2,3,4]:
                    if f[i]%(f[5]/f[6])!=0:
                        s="Klasse %s, Fach %s:  '%s' sollte ein x-faches von "%(f[0],f[1],columns[i])
                        s+="'Blockstunden'/'In Klassengruppes geteilt' "
                        s+="(Hier: %.1f > (%.1f/%.1f) *%d)"%(f[i],f[5],f[6],f[i]//(f[5]/f[6]))
                        self.message.append(s)
            
    def corresponding_faecher(self,fach,klasse,worksheet=""):

        """find the fach names corresponding to "fach" in a given class"""

        if fach in self.fach_list:

            return [fach]

        temp=[i for i in self.faecher_raw[self.klassen_dict[klasse]] if i[1]==fach]

        if len(temp)==0:

            self.message.append("Warning: in %s, klasse %s, fach '%s' existiert nicht"%(worksheet,klasse,fach))

            return []

        res =["%s__G%d_T%d_D%d"%f for f in separate_faecher(self.hauptfaecher,temp)]

        return res

    def get_ubergreifend(self,fach,klasse):
        """ Gibt den Anzahl von Klassen, die zusammen ein fach haben"""
        res=1
        klasse=self.klassen_list[klasse]
        fach=self.fach_list[fach]
        l=[j for j in self.ubergreifend if j[0]==fach and klasse in j[2]]
        for i in l:
            res=len(i[2])
        return res

    def check_raume(self):
        """ Prüft, dass die Räume lang genug verfugbar sind, 
        so dass alle lehrveranstaltungen statt finden können"""
        for r in range(len(self.raum_verfugbar)):
            M = np.sum(self.raum_verfugbar[r])
            N = np.sum([[ self.stunden[k][i]*self.geteilt[i]/self.get_ubergreifend(i,k) 
                    for i,f in enumerate(self.fach_list) 
                         if f in self.raum_faecher[r]]
                    for k in range(len(self.klassen_list))])
            if N>M:
                s="Raum nummer %d ist nur %d Stunde verfugbar, "%(r+1,M)
                s+="aber es sollte %.1f Stunden dort gelehrt werden"%(N)
                self.message.append(s)

        
    def import_faecher(self):

        # first, determine which courses are to be counted as "Hauptfach" (main course)

        lehrerfaecher=[]

        for row in self.iter_row_from(sheet="Lehrerübersicht", 

                                      start="Kann ...unterrichten",

                                      cols=["Kann ...unterrichten"]):

            lehrerfaecher.append(to_list(row[0]))

        hauptfaecher=set.intersection(*[set(i) for i in lehrerfaecher 

                                        if "Deutsch" in i])

        # exclude ubergreifende faecher

        ubergreifend=[]

        for row in self.iter_row_from("Übergreifende Fächer","Fach"):

            if not is_none(row[0]):

                ubergreifend.append(to_str(row[0]))

        hauptfaecher=hauptfaecher.difference(set(ubergreifend))

        # exclude Mittagspause

        keinhauptfach=["Mittagspause"]

        hauptfaecher=hauptfaecher.difference(set(keinhauptfach))       

        

        # exclude gleichzeitige faecher

        sheet="Gleichzeitige Fächer"

        gleichzeitig=[]

        for row in self.iter_row_from(sheet,"Fächer"):

            if not is_none(row[0]):

                temp=[to_str(i) for i in row if not is_none(i)]

                gleichzeitig.extend(temp)

        hauptfaecher=hauptfaecher.difference(set(gleichzeitig))

        

        self.hauptfaecher=hauptfaecher

        # second, read "Lehrpläne" to get the course´s properties

        # if necessary, create 2 courses instead of one to account for the course 

        # properties (if it needs a tandemlehrer occasionnally for instance)

        faecher=[]

        faecher_raw=[]

        fach_set=[]

        current=""

        temp=[]

        raw=self.iter_row_from("Lehrpläne","Klasse")

        for row_id,row in enumerate(raw):

            if not is_none(row[0]):

                # here we get to a new class

                

                # call to 'separate_faecher' to:

                # -> regoup the hauptfaecher in one course

                # -> create multiple courses instead of one,

                #       in order to account for all of the course properties

                if current!="":

                    faecher_raw.append(temp)

                    fach_set.extend([i[1] for i in temp])

                    to_extend=separate_faecher(self.hauptfaecher,temp)

                    faecher.extend(to_extend)

                current=to_str(row[0])

                temp=[]

                

            if not is_none(row[1]) and current!="":

                # fetch the data in correct format

                temp.append((current,

                    to_str(row[1]),

                    to_float(row[2]),

                    to_float(row[3]),

                    to_float(row[4]),

                    to_int(row[5],default=1),

                    to_int(row[6],default=1),

                    to_str(row[7],default="muss")))

                

                if to_str(row[7],"muss") not in ["kann","muss"]:

                    # create warning message

                    self.message.append(

                        "Warning in Lehrpläne, row %d: "%row_id+

                        "'Muss geteilt werden?' ist entweder "+

                        "'kann' oder 'muss'")

                    row[7].value=None

        faecher_raw.append(temp)

        to_extend=separate_faecher(self.hauptfaecher,temp)

        faecher.extend(to_extend)

        

        # finally list all found faecher, give them name and store them          

        faecher=list(set(faecher))

        temp=[]

        for i in set([i[0] for i in faecher]):

            l = [j for j in faecher if j[0]==i]

            if len(l)==1:

                temp.append(l[0])

            else:

                "Create unique name for each fach"

                for fach in l:

                    temp.append(("%s__G%d_T%d_D%d"%fach,*fach[1:]))

        faecher=sorted(temp,key=lambda x: x[0])

        faecher.append(("Tandem",1,0,1))

        

        self.faecher_raw = faecher_raw

        self.fach_set    = set(fach_set)

        self.fach_list   = [i[0] for i in faecher]

        self.geteilt     = [int(i[1]) for i in faecher]

        self.tandem      = [i[2] for i in faecher]

        self.dauer       = [i[3] for i in faecher]

        self.n_fach      = len(faecher)

        

        self.fach_dict=dict(zip(self.fach_list,

                                range(len(self.fach_list))))

    def import_faecherstunden(self):

        fraw = sorted(self.faecher_raw,key=lambda x: x[0][0])

        klassen_list = [i[0][0] for i in fraw]

        

        stunden = []

        gleichgultig = []

        klassenfach = []

        

        for faecher in fraw:

            klasse=faecher[0][0]

            temp=[0]*len(self.fach_list)

            temp_eq=[]

            temp_klassenfach=["Tandem"]

            for f in faecher:

                name = "Hauptfach" if f[1] in self.hauptfaecher else f[1]

                if name in self.fach_list:

                    temp[self.fach_dict[name]]+=f[2]

                    temp_klassenfach.append(name)

                else:

                    sf=separate_faecher(self.hauptfaecher,[f])

                    sf=[("%s__G%d_T%d_D%d"%i,*i[1:]) for i in sf]

                    

                    

                    for n_tandem in range(3):

                        stun = f[2+n_tandem]-f[3+n_tandem] if n_tandem<2 else f[2+n_tandem]

                        if stun<0:

                            self.message.append(

                            "Im 'Lehrepläne', klasse '%s', fach '%s',"%(klasse,f[1])+

                            " es kann nicht mehr Tandem als Lehrer geben (diff=%f)"%stun)

                        # sf2: list of equivalent Fächer with properties

                        if f[7]=="kann":

                            sf2=[j[0] for j in sf if j[1]*(1+j[2])==n_tandem+1]

                            if len(sf2)==1:

                                temp[self.fach_dict[sf2[0]]]+=stun

                            elif len(sf2)>1:

                                sf2=set(sf2)

                                add_gleichgultig(temp_eq,sf2,stun)

                        else:

                            sf2=[j[0] for j in sf if j[2]==n_tandem]

                            if len(sf2)==1:

                                temp[self.fach_dict[sf2[0]]]+=stun

                    

                    for i in sf:

                        temp_klassenfach.append(i[0])

            

            klassenfach.append(sorted(set(temp_klassenfach)))

            stunden.append((klasse,temp))

            gleichgultig.append((klasse,temp_eq))

        

        temp=zip(stunden,gleichgultig,klassenfach)

        temp=sorted(temp,key=lambda x: x[0][0])

        

        self.klassen_list = [i[0][0] for i in temp]

        self.klassenfach=[i[2] for i in temp]

        self.stunden=[i[0][1] for i in temp]

        self.gleichgultig=[i[1][1] for i in temp]

        self.klassen_dict = dict(zip(self.klassen_list,

                                     range(len(self.klassen_list))))

    def import_lehrer(self):

        sheet="Lehrerübersicht"

        name=[]

        

        # get raw data in excel sheet

        for row in self.iter_row_from(sheet,"Name"):

            if not is_none(row[0]):

                name.append( (to_str(row[0]),

                             to_str(row[1]),

                             to_str(row[2]),

                             None,

                             to_int(row[4]),

                             to_int(row[5]),

                             to_int(row[6]),

                             to_list(row[7]),

                             to_int(row[8])) )

        

        name=sorted(name,key=lambda x: x[2])

        lehrer_faecher = []

        for lehrer in name:

            temp=[]

            if lehrer[5]==1:

                self.add_fach(temp,["Sport","SportW"])

            if lehrer[6]==1:

                self.add_fach(temp,["Sport","SportM"])

            if lehrer[8]==1:

                self.add_fach(temp,["Tandem"])

            

            self.add_fach(temp,lehrer[7])

            temp=list(set(temp))

            lehrer_faecher.append(temp)

        

        

        lehrer_dict={}

        for i in enumerate(name):

            lehrer_dict[i[1][2]]=i[0]

        

        self.lehrer_list    = [i[2] for i in name]

        self.lehrer_dict    = lehrer_dict

        self.arbeitzeit     = [i[4] for i in name]

        self.lehrer_name    = [i[0] for i in name]

        self.lehrer_vorname = [i[1] for i in name]

        self.lehrer_faecher = lehrer_faecher

    def import_klassen_tage(self):

        klassen_tage={}

        for row in self.iter_row_from("Stundenzahl pro Tag","Klasse"):

            klasse = to_str(row[0])

            if klasse not in self.klassen_list:

                self.message.append("Warning: in 'Stundenzahl pro Tag', "+

                                "Klasse '%s' ist unbekannt."%klasse)

            else:

                klassen_tage[klasse]=[to_int(i) for i in row[1:6]]

        

        error = set(self.klassen_list).difference(set(klassen_tage.keys()))

        if len(error)>0:

            self.message.append("Warning: in 'Stundenzahl pro Tag', "+

                "Klasses %r are missing"%error)

            for i in error:

                klassen_tag[i]=[0,0,0,0,0]

            

        woche=[int(i) for i in np.max(list(klassen_tage.values()),0)]

            

        klassen_zeiten={}

        klassen_tagende={}

        for key,tage in klassen_tage.items():

            current=0

            temp=[]

            temp2=[]

            for i in range(5):

                temp.extend([1]*tage[i])

                temp.extend([0]*(woche[i]-tage[i]))

                temp2.append(tage[i]+current)

                current+=woche[i]

            klassen_zeiten[key]=temp

            klassen_tagende[key]=temp2

        anfang=[1]
        anfang.extend(woche)
        anfang=[int(i) for i in np.cumsum(anfang)[:-1]]
        

        self.klassen_tage=klassen_tage

        self.klassen_tagende=klassen_tagende            

        self.klassen_zeiten=klassen_zeiten

        self.woche=woche

        self.tag_anfang= anfang
        
        self.stunde_pro_tag=max(self.woche)

        self.n_zeitslots=sum(self.woche)

    def import_klassen_lehrer(self):

        klassen_lehrer={}

        klassen_tandem={}

        klassen_partner={}

        for row in self.iter_row_from("Klassenlehrer","Klasse"):

            klasse = to_str(row[0])

            if klasse == "":

                continue

            if klasse not in self.klassen_list:

                self.message.append("Warning: im 'Klassenlehrer', "+

                                "Klasse '%s' ist unbekannt."%klasse)

            else:

                klassen_lehrer[klasse]=to_str(row[1])

                klassen_tandem[klasse]=to_str(row[2])

                klassen_partner[klasse]=to_list(row[3])

        

        self.check_all_klassen(klassen_lehrer.keys(),sheet="Klassenlehrer")

        self.check_in_lehrer(klassen_tandem.values(),sheet="Klassenlehrer")

        self.check_in_lehrer(klassen_lehrer.values(),sheet="Klassenlehrer")

        for val in klassen_partner.values():
            self.check_in_lehrer(val,sheet="Klassenlehrer")

        self.klassen_lehrer=klassen_lehrer

        self.klassen_tandem=klassen_tandem

        self.klassen_partner=klassen_partner

    def import_raume(self):

        """ Read "Räume" sheet to know the room availabilities

        """

        raum_faecher=[]

        raum_verfugbar=[]

        current_faecher=[]

        current_verfugbar=[]

        raw=self.iter_row_from("Räume","Fächer",offset=(0,-1))

        raw.append(["end"])

        for row in raw:

        

            #first append the previous buffers

            temp_bool=row[0]=="end"

            temp_bool= temp_bool if temp_bool else to_str(row[0])=="Faecher"

            if temp_bool:

                if len(current_faecher)>0:

                    raum_faecher.append(current_faecher)

                    raum_verfugbar.append(self.check_zeiten(

                        current_verfugbar,"Räume"))

                if row[0]=="end":

                    break

                # create new buffers

                current_faecher = []

                current_verfugbar = []

                temp=[to_str(i) for i in row[1:] if not is_none(i)]

                bad,good=self.check_in_faecher(temp,"Räume")

                self.add_fach(current_faecher,good)

            else:

                # insert data in the buffers

                stunde = to_int(row[0],default=None)

                if stunde is not None:

                    current_verfugbar.append([to_int(i) for i in row[1:6]])

        self.raum_faecher=raum_faecher

        self.raum_verfugbar = raum_verfugbar

        self.n_raum = len(raum_faecher)

    def import_ubergreifend(self):

        ubergreifend=[]

        sheet="Übergreifende Fächer"

        for row in self.iter_row_from(sheet,"Fach"):

            if not is_none(row[0]):

                faecher=[]

                self.add_fach(faecher,[to_str(row[0])])

                err,klassen=self.check_in_klassen(to_list(row[1]),sheet)

                for fach in faecher:

                    ubergreifend.append((

                        fach,

                        klassen))

        faecher = list(set([i[0] for i in ubergreifend]))

        for f in faecher:

            temp=[]

            for k in [i[1] for i in ubergreifend if i[0]==f]:

                temp.extend(k)

            other_klassen=self.klassen_set.difference(temp)

            for k in other_klassen:

                ubergreifend.append((f,[k]))

        res=[]

        for f in sorted(faecher):

            temp=[i for i in ubergreifend if i[0]==f]

            temp=[(i[1][0],i[0]+1,i[1][1]) for i in enumerate(temp)]

            res.extend(temp)

        self.ubergreifend=res

    def import_vorgaben(self):

        sheet="Feste Vorgaben"

        vorgaben = {}

        vorgaben_zeiten ={}

        vorgaben_mit_lehrer = {}

        vorgaben_ml_zeiten ={}

        vorgaben_max = 0

        cur=""

        temp=[]

        fach_dict={}

        raw=self.iter_row_from(sheet,"Montag",offset=(-1,-1))

        raw.append("end")

        for row in raw:

            # first treat the previously read data

            if not is_none(cur) and (row=="end" or to_str(row[1])=="Montag"):

                temp=self.check_zeiten(temp,default=())

                fach_dict={}

                for z,vor in enumerate(temp):

                    z=z+1

                    for i in vor:

                        faecher=self.corresponding_faecher(i[0],cur,sheet)

                        index=[cur]

                        try:

                            index.append(fach_dict[i[0]])

                        except KeyError:

                            length=len(fach_dict)+1

                            fach_dict[i[0]]=length

                            index.append(length)

                            vorgaben_max=max(vorgaben_max,length)

                        index=tuple(index)

                        

                        if i[1] is None:

                            if index not in vorgaben.keys():

                                vorgaben_zeiten[index]=[]

                                vorgaben[index]=faecher

                            vorgaben_zeiten[index].append(z)

                        else:

                            for lehrer in i[1]:

                                index2=(index[0],lehrer,index[1])

                                if index2 not in vorgaben.keys():

                                    vorgaben_ml_zeiten[index2]=[]

                                    vorgaben_mit_lehrer[index2]=faecher

                                vorgaben_ml_zeiten[index2].append(z)

            if row=="end":

                break

            if to_str(row[1])=="Montag":

                temp = []

                bad,good=self.check_in_klassen([to_str(row[0])],"Fest Vorgaben")

                cur= good[0] if len(good)>0 else ""

            elif not is_none(row[0]):

                temp.append([to_list_2(i) for i in row[1:6]])

                
        self.vorgaben_zeiten=vorgaben_zeiten

        self.vorgaben_ml_zeiten=vorgaben_ml_zeiten

        self.vorgaben=vorgaben

        self.vorgaben_mit_lehrer=vorgaben_mit_lehrer

        self.vorgaben_max=vorgaben_max

    def import_lehrerverfugbarkeiten(self):

        verfug=[]

        res={}

        sheet="Lehrerverfügbarkeiten"

        for row in self.iter_row_from(sheet, "Stunde"):

            if is_none(row[0]):

                continue

            temp=[]

            for cell in row[1:6]:

                lehrer=self.check_in_lehrer(to_list(cell),sheet)[1]

                temp.append(lehrer)

            verfug.append(temp)

        verfug=self.check_zeiten(verfug)

        for z,lehrer_list in enumerate(verfug):

            for lehrer in lehrer_list:

                try:

                    res[lehrer].append(z)

                except KeyError:

                    res[lehrer]=[z]

        for i,vals in res.items():

            vals=[i+1 for i in range(self.n_zeitslots) 

                  if i not in vals]

            res[i]=vals

                    

        self.lehrer_verfugbar=res

    def import_gleichzeitig(self):

        sheet="Gleichzeitige Fächer"

        gleichzeitig=[]

        gleichzeitig_geteilt=[]

        

        for row in self.iter_row_from(sheet,"Fächer"):

            if not is_none(row[0]):

                temp=[]

                vals=[to_str(i) for i in row if not is_none(i)]

                bad,vals=self.check_in_faecher(vals,sheet)

                if len(vals)==1:

                    self.add_fach(temp,vals,check_geteilt=">1")

                    gleichzeitig_geteilt.extend(temp)

                elif len(vals)>1:

                    self.add_fach(temp,vals,check_geteilt=len(vals))

                    for i in range(len(temp)-1):

                        gleichzeitig.append((temp[i],temp[i+1]))

        self.gleichzeitig = gleichzeitig

        self.gleichzeitig_geteilt = gleichzeitig_geteilt

    def write_file(self):

        res = [('nLehr' , len(self.lehrer_list))]

        res.append(('Lehrer' , write_list(self.lehrer_list) ))

        res.append(('LehrerFaecher',write_list_list(self.lehrer_faecher)))

        res.append(('ArbeitZeit',write_list(self.arbeitzeit)))

        res.append(('LehrerVerfugbar',write_dict(self.lehrer_verfugbar)))

        res.append(('nKlas' , len(self.klassen_list) ))

        res.append(('Klassen' , write_list(self.klassen_list)))

        res.append(('KlassenLehrer',write_dict(self.klassen_lehrer)))

        res.append(('TandemLehrer',write_dict(self.klassen_tandem)))

        res.append(('PartnerLehrer',write_dict(self.klassen_partner)))

        res.append(('UnterrichtStunden',write_list_list(self.stunden,bracket=False) ))

        res.append(('nFach' , self.n_fach ))

        res.append(('Faecher' , write_list(self.fach_list)))

        res.append(('FachDauer',write_list(self.dauer)))

        res.append(('TandemNummer',write_list(self.tandem)))

        res.append(('KlassenFaecher',write_list_list(self.klassenfach)))

        res.append(('Gleichgultig',write_list3([

            [i[0] for i in j] for j in self.gleichgultig])))

            
        n = max([len(i) for i in self.gleichgultig])
        
        res.append(("nGGStunden", n))
        temp=[ [*[ i[1] for i in j],*([0.0]*(n-len(j)))] for j in self.gleichgultig]
        res.append(('GleichgultigStunden',write_list_list(temp, bracket=False)))

        temp=dict([((i[0],i[1]),i[2]) for i in self.ubergreifend])

        res.append(('nUbergreifendMax',max([i[1] for i in self.ubergreifend])))

        res.append(('Ubergreifend',write_dict(temp,tuple_key=2)))

        res.append(('GeteilteFach',write_list(self.geteilt)))

        res.append(('GleichzeitigGeteilteFach',write_list(self.gleichzeitig_geteilt)))

        temp=" \n".join(["('%s') '%s' "%i for i in self.gleichzeitig])

        res.append(('GleichzeitigFach',"[%s]"%temp))

        res.append(('nZeit', self.n_zeitslots))

        res.append(('KlassenTage' , write_dict(self.klassen_tage)))

        res.append(('KlassenTageEnde' , write_dict(self.klassen_tagende)))

        res.append(('KlassenZeiten' , write_list_list(

            [self.klassen_zeiten[i] for i in self.klassen_list],bracket=False)))

        res.append(('Woche',write_list(self.woche)))
        
        res.append(("TagAnfang",write_list(self.tag_anfang)))

        res.append(('nRaum',self.n_raum))

        res.append(('RaumFaecher',write_list_list(self.raum_faecher)))

        res.append(('RaumVerfugbar',write_list_list(self.raum_verfugbar,bracket=False)))

        res.append(('nVorgabenMax',max(1,self.vorgaben_max)))

        res.append(('Vorgaben',write_dict(self.vorgaben,tuple_key=2)))

        res.append(('VorgabenZeiten',write_dict(self.vorgaben_zeiten,tuple_key=2)))

        res.append(('VorgabenMitLehrer',write_dict(self.vorgaben_mit_lehrer,

                                                     tuple_key=3)))

        res.append(('VorgabenMLZeiten',write_dict(self.vorgaben_ml_zeiten,

                                                     tuple_key=3)))

        res=["'%s' : %s"%i for i in res]

        with open(self.target_file,"w") as file:

            file.write("\n \n ".join(res))  

     

# main function
# getting arguments for the source and destination file7

def main(write_files=True,source=r"school_data.xlsx",destination=r"school_requirements.dat",command=False):

    """

    Get arguments, create a reader object and write a .dat file and .message file

    """

    # reading the command line arguments

    if command:

        parser = argparse.ArgumentParser(description='Get source and destination file')

        parser.add_argument("--source", dest="source", type=str, 

                            default=source,

                            help='The source data file (excel format)')

        parser.add_argument('--dest', dest='destination', type=str,

                            default=destination,

                            help='The destination file (mosel .dat format)')

        args = parser.parse_args()

        

        source=args.source

        destination=args.destination

        write_files=True

    

    data_excel = os.path.join(folder,source)

    target_file = os.path.join(folder,destination)

    data = load_workbook(data_excel)

       

	

    r=reader(source=data_excel,target_file=target_file)

    

    if write_files:

        r.write_file()

        s="\n".join(r.message)

        with open(os.path.join(folder,"data_loader_message.txt"),"w") as f:

            f.write("### Warnings from the data processing ###\n")

            f.write("           (better if empty!)\n\n")

            f.write(s)

    return r

if __name__=="__main__":

    main(command=True)

# Remarks: 

#  - gleichgultig is empty

